"""Excel export for BOM with multipliers and evidence."""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List, Dict


class ExcelExporter:
    """Export BOM to Excel with formatting and evidence."""
    
    def export(self, bom_items: List, hierarchy_data: Dict, 
              output_path: str):
        """Export BOM to Excel file."""
        wb = Workbook()
        
        # Remove default sheet
        wb.remove(wb.active)
        
        # Create BOM sheet
        bom_sheet = self._create_bom_sheet(wb, bom_items)
        
        # Create hierarchy sheet
        hierarchy_sheet = self._create_hierarchy_sheet(wb, hierarchy_data)
        
        # Create evidence sheet
        evidence_sheet = self._create_evidence_sheet(wb, bom_items)
        
        wb.save(output_path)
    
    def _create_bom_sheet(self, wb: Workbook, bom_items: List) -> None:
        """Create main BOM sheet."""
        ws = wb.create_sheet("BOM", 0)
        
        # Headers
        headers = [
            "Item ID", "Category", "Subtype", "Thread", "Length", 
            "Head Type", "Material", "Brand", "Per Instance", 
            "Multiplier", "Total Count", "Confidence", "Needs Review", "Reasoning"
        ]
        ws.append(headers)
        
        # Style header row
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Add data rows
        for item in bom_items:
            row = [
                item.item_id,
                item.category,
                item.subtype,
                item.spec_normalized.get("thread", ""),
                item.spec_normalized.get("length", ""),
                item.spec_normalized.get("head_type", ""),
                item.spec_normalized.get("material", ""),
                item.spec_normalized.get("brand_code", ""),
                item.per_instance_count,
                item.multiplier,
                item.total_count,
                f"{item.confidence:.2f}",
                "Yes" if item.needs_review else "No",
                item.reasoning
            ]
            ws.append(row)
            
            # Highlight rows needing review
            if item.needs_review:
                row_num = ws.max_row
                for col in range(1, len(headers) + 1):
                    cell = ws.cell(row=row_num, column=col)
                    cell.fill = PatternFill(
                        start_color="FFF2CC", 
                        end_color="FFF2CC", 
                        fill_type="solid"
                    )
        
        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
        
        # Add summary row
        summary_row = ws.max_row + 2
        ws.cell(row=summary_row, column=1, value="Summary:")
        ws.cell(row=summary_row, column=1).font = Font(bold=True)
        
        ws.cell(row=summary_row + 1, column=1, value="Total Items:")
        ws.cell(row=summary_row + 1, column=2, value=len(bom_items))
        
        ws.cell(row=summary_row + 2, column=1, value="Items Needing Review:")
        ws.cell(row=summary_row + 2, column=2, value=sum(1 for item in bom_items if item.needs_review))
        
        ws.cell(row=summary_row + 3, column=1, value="Total Fasteners:")
        ws.cell(row=summary_row + 3, column=2, value=sum(item.total_count for item in bom_items))
    
    def _create_hierarchy_sheet(self, wb: Workbook, hierarchy_data: Dict) -> None:
        """Create hierarchy visualization sheet."""
        ws = wb.create_sheet("Hierarchy")
        
        # Headers
        headers = ["Node ID", "Page", "Title", "Type", "Multiplier", "Parent", "Children"]
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Add hierarchy nodes
        for node in hierarchy_data["hierarchy"]:
            row = [
                node.node_id,
                node.page_num + 1,
                node.title,
                node.node_type,
                node.multiplier,
                node.parent_id or "",
                ", ".join(node.children) if node.children else ""
            ]
            ws.append(row)
        
        # Auto-adjust widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)
    
    def _create_evidence_sheet(self, wb: Workbook, bom_items: List) -> None:
        """Create evidence tracking sheet."""
        ws = wb.create_sheet("Evidence")
        
        # Headers
        headers = ["Item ID", "Source Nodes", "Detection Count", "Evidence IDs"]
        ws.append(headers)
        
        # Style header
        header_fill = PatternFill(start_color="C55A11", end_color="C55A11", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
        
        # Add evidence rows
        for item in bom_items:
            row = [
                item.item_id,
                ", ".join(item.source_nodes),
                item.per_instance_count,
                ", ".join(item.evidence_detections[:10])  # Limit to first 10
            ]
            ws.append(row)
        
        # Auto-adjust widths
        for col in range(1, len(headers) + 1):
            max_length = 0
            column = get_column_letter(col)
            for cell in ws[column]:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            ws.column_dimensions[column].width = min(max_length + 2, 50)

